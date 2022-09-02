#-*- coding:utf-8 -*-
# Create by  zirui.wang
# Create on 2020/11/3
import openpyxl,os,json
class open_execl(object):
    def get_excel(self):
        '''打开xlsx格式的excel,可指定打开某个xlsx在B=位置输入‘xxxx.xlsx’ '''
        b='chengshi.xlsx'
        global a11#xlsx文件（excel）全局变量
        a11=os.path.dirname(os.getcwd())+'/confige/'+b
        get_excel=openpyxl.load_workbook(a11)
        return get_excel
    def get_sheetnames(self):
        '''获取所有shee页名'''
        sheetnames=self.get_excel().sheetnames
        return sheetnames
    def get_sheetname(self,suoyin=None):
        '''获取一页的所有内容。索引为空时默认为0打开第一页。也可指定打开某一页输入0 or 1 or 2'''
        if suoyin==None:
            sheetname=self.get_excel()[self.get_sheetnames()[0]]
            return sheetname
        else:
            sheetname1=self.get_excel()[self.get_sheetnames()[suoyin]]
            return sheetname1
    def get_values(self,row,col,suoyin=None):
        '''获取单一单元格的values内容'''
        date=self.get_sheetname(suoyin).cell(row,col).value
        return date
    def max_row(self,suoyin=None):
        '''获取excel最大行数'''
        max_row=self.get_sheetname(suoyin).max_row
        return max_row
    def max_column(self,suoyin=None):
        '''获取最大列数'''
        max_col = self.get_sheetname(suoyin).max_column
        return max_col
    def row_value(self,suoyin1=None,suoyin=None):
        '''取出一行内容,suoyin1为空时默认为2(第2行）'''
        if suoyin1==None:
            suoyin1=2
        row_value=self.get_sheetname(suoyin)[suoyin1]
        a=[]
        for i in row_value:

            a.append(i.value)
        json.dumps(a)

        return a
    def wj_xr(self,row,col,value):
        '''运行结果写入excel'''
        wd=self.get_excel()#打开excel
        wr=wd.active#激活
        wr.cell(row,col,value)#写入
        wd.save(a11)#保存excel
    # def fs_youjian(self,mubiaoEmail,zhuti,fujianmingzi,neirong=None):
    #     '''发送邮件'''
    #     a=os.path.dirname(os.getcwd())+'/confige/'+fujianmingzi
    #     yag=yagmail.SMTP(user='18701219401@163.com',password='GQWOTFABQHOJIBPE',host='smtp.163.com')
    #     yag.send(to=[mubiaoEmail],subject=zhuti,contents=[a,neirong])

if __name__=='__main__':
    open=open_execl()
    print(open.get_excel())
    print(open.get_sheetnames())
    print(open.get_sheetname())
    print(open.get_values(2,3))
    print(open.max_row())
    print(open.max_column())
    print(open.row_value())
