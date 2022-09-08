import xlrd2
import cpca
import pandas as pd

import xlwt
import openpyxl


class Employee():
    def __init__(self, content, sheng, shi, xian):
        self.content = content
        self.sheng = sheng
        self.shi = shi
        self.xian = xian

    def print_hi():
        # xlsx = xlrd2.open_workbook('D:/合约商机线上数据-给到振坤.xlsx')
        xlsx = xlrd2.open_workbook('D:/A-全国物业注册地址-9.2.xlsx')
        #  ExcelReader reader2 = ExcelUtil.getReader("D:/合约商机线上数据-给到振坤1.xlsx", 2);
        # 通过sheet名查找：xlsx.sheet_by_name("sheet1")
        # 通过索引查找：xlsx.sheet_by_index(3)
        table = xlsx.sheet_by_index(0)

        # 获取单个表格值 (2,1)表示获取第3行第2列单元格的值
        # value = table.cell_value(2, 1)
        # print("第3行2列值为", value)

        # 获取表格行数
        nrows = table.nrows
        print("表格一共有", nrows, "行")
        list1 = []
        # 获取第1列所有值（列表生成式）
        name_list = [str(table.cell_value(i, 0)) for i in range(1, nrows)]
        # print("第4列所有的值：", name_list)
        for i in name_list:
            # print("序号：%s   值：%s" % (name_list.index(i) + 1, i))
            try:
                df = cpca.transform_text_with_addrs(i)
            except KeyError as e:
                df = None
            if df is None:
                # print("序号：%s   值：%s " % (name_list.index(i) + 1, i))
                emp1 = Employee(i, "null", "null", "null")
                list1.append(emp1)
            else:
                # print("序号：%s   值：%s 分词结果：%s" % (name_list.index(i) + 1, i, df))
                a = df.to_dict('dict')
                emp1 = Employee(i, list(a.get("省").values()), list(a.get("市").values()), list(a.get("区").values()))
                list1.append(emp1)
        # 创建sheet工作表
        # sheet1 = file.add_sheet('sheet1', cell_overwrite_ok=True)
        outwb = openpyxl.Workbook()
        outws = outwb.create_sheet(index=0)
        # outws.cell(row, col).value = row * 2
        # 先填标题
        # sheet1.write(a,b,c) 函数中参数a、b、c分别对应行数、列数、单元格内容
        outws.cell(1, 1).value = "内容"  # 第1行第1列
        outws.cell(1, 2).value = "省"  # 第1行第2列
        outws.cell(1, 3).value = "市"  # 第1行第3列
        outws.cell(1, 4).value = "区"  # 第1行第4列
        # 循环填入数据
        for i in range(len(list1)):
            outws.cell(i + 2, 1).value = list1[i].content  # 第1列
            outws.cell(i + 2, 2).value = str(list1[i].sheng[0])  # 第2列
            outws.cell(i + 2, 3).value = str(list1[i].shi[0])  # 第3列
            outws.cell(i + 2, 4).value = str(list1[i].xian[0])  # 第4列
        outwb.save('A-全国物业注册地址-9.2结果1.xls')


if __name__ == '__main__':
    df = cpca.transform_text_with_addrs("广东省深圳市龙华区")
    print(df)
    # Employee.print_hi()
